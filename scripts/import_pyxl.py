#!/usr/bin/env python
# -*- coding: utf-8 -*-

workbook_name = '/home/mario/Dropbox/SharedWithMario/PvdT/DenHelder_for_Ghini.xlsx'
sheet_name = 'Den Helder'

column_equivalence = {
    'remarks': '',
    'IUCN status': '',
    'Accessionnumber': '',
    'Type of material': '',
    'Family on label': '',
    'Genus on label': '',
    'Hybrid marker': '',
    'Specific epithet': '',
    'Auteur 1': '',
    'Auteur 2': '',
    'Infraspecific Rank': '',
    'Infraspecific epithet': '',
    'Cultivar': '',
    'Synonyms': '',
    'Database used': '',
    'Checked with Tropicos': '',
    'Name status': '',
    'Distribution CJB': '',
    'Distribution on label': '',
    'Distribution print': '',
    'Habitat': '',
    'Provenance Code': '',
    'Provenance (breeder/botanical garden)': '',
    'Provenance (finding place)': '',
    'Plantation date': '',
    'Flowering period': '',
    'Flower colour': '',
    'plant beschrijving': '',
    'speciale uiterlijke kenmerken': '',
    'Dutch name': '',
    'Other names': '',
    'Garden Acronym': '',
    'Location in Oranjerie': '',
    'Location in garden': '',
    'Number of plants': '',
    'Location specification': '',
    'Plant status': '',
    'Sticker size': '',
    'Print ready': '',
    'Print aantal': '',
    'Gebruik': '',
    'Determinated by': '',
    'Namecheck by': '',
    'Namecheck date': '',
    'Literatuur': '',
    'Opmerkingen': '',
    'foto': '',
    }

import openpyxl
wb = openpyxl.load_workbook(workbook_name)
sheet = wb.get_sheet_by_name(sheet_name)
max_column = max(i for i in range(sheet.max_column) if sheet.cell(row=1, column=i+1).value is not None) + 1
header = [sheet.cell(row=1, column=i+1).value for i in range(max_column)]

for rn in range(1, sheet.max_row):
    row = dict(list(zip(header, [sheet.cell(row=rn+1, column=i+1).value for i in range(max_column)])))
